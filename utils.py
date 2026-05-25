
import json
import logging
import os
import sys
import time
from typing import Optional, Type, TypeVar

import streamlit as st
from langchain_google_genai import ChatGoogleGenerativeAI
from pydantic import BaseModel, ValidationError
from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential

from json_repair_utils import parse_json_resilient

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stdout,
)


def validate_api_key() -> bool:
    """Check that GOOGLE_API_KEY is present in the environment."""
    key = os.getenv("GOOGLE_API_KEY")
    return bool(key and key.strip())


def show_api_key_error() -> None:
    """Display a friendly error when the API key is missing and halt the app."""
    st.error(
        "⚠️ **Google API Key not found.**\n\n"
        "Please set your `GOOGLE_API_KEY` environment variable.\n\n"
        "**Quick fix:**\n"
        "1. Create a `.env` file in the project root\n"
        '2. Add:  `GOOGLE_API_KEY=your_actual_key_here`\n'
        "3. Restart the app\n\n"
        "[Get a key →](https://aistudio.google.com/app/apikey)",
        icon="🔑",
    )
    st.stop()


def _is_retryable(exc: Exception) -> bool:
    """Return True for transient errors that merit a retry."""
    # Network-level errors
    if isinstance(exc, (ConnectionError, TimeoutError)):
        return True
    # Google GenAI API errors
    try:
        from google.genai.errors import ServerError, ClientError
        if isinstance(exc, ServerError):
            return True
        if isinstance(exc, ClientError) and getattr(exc, "code", None) == 429:
            return True
    except Exception:
        pass
    return False


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception(_is_retryable),
    reraise=True,
)
def invoke_with_retry(chain, inputs: dict):
    """Invoke an LCEL chain with automatic retry on transient API failures."""
    return chain.invoke(inputs)


from json_repair import repair_json


def repair_llm_json(text: str) -> str:
    """Strip markdown fences and repair common LLM JSON errors."""
    text = text.strip()
    if text.startswith("```json"):
        text = text[7:]
    elif text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()
    return repair_json(text)


@st.cache_resource
def load_model() -> ChatGoogleGenerativeAI:
    """Load and cache the LLM."""
    return ChatGoogleGenerativeAI(model="gemini-2.5-flash-lite", temperature=0.3)


def get_score_class(score: int) -> str:
    """Return CSS class for quality health score colouring."""
    if score >= 75:
        return "score-good"
    elif score >= 50:
        return "score-warning"
    return "score-risk"


def get_severity_badge_class(severity: str) -> str:
    """Map severity string to CSS badge class."""
    from constants import SEVERITY_BADGE_MAP
    return SEVERITY_BADGE_MAP.get(severity.lower(), "badge-medium")


def build_steps_html(steps: list[str]) -> str:
    """Convert a list of steps into styled HTML."""
    return "".join([
        f"<div style='margin-bottom:0.3rem;'>"
        f"<span style='color:#3b82f6;font-family:Space Mono,monospace;margin-right:0.5rem;'>{i + 1}.</span>"
        f"{s}</div>"
        for i, s in enumerate(steps)
    ])


def build_list_html(items: list[str], icon: str, color: str) -> str:
    """Build a styled HTML list from string items."""
    return "".join([
        f"<div style='padding:0.35rem 0;border-bottom:1px solid #1a2535;font-size:0.9rem;'>"
        f"<span style='color:{color};margin-right:0.5rem;'>{icon}</span>"
        f"{item}</div>"
        for item in items
    ])


def get_category_badge_class(category: str) -> str:
    """Map category string to CSS badge class."""
    from constants import CATEGORY_BADGE_MAP
    return CATEGORY_BADGE_MAP.get(category.lower(), "badge-medium")


def get_priority_badge_class(priority: str) -> str:
    """Map priority string to CSS badge class."""
    from constants import PRIORITY_BADGE_MAP
    return PRIORITY_BADGE_MAP.get(priority.lower(), "badge-medium")


def get_auto_feasibility_class(feasibility: str) -> str:
    """Map automation feasibility to CSS badge class."""
    from constants import AUTO_FEASIBILITY_MAP
    return AUTO_FEASIBILITY_MAP.get(feasibility.lower(), "badge-medium")


def get_auto_effort_class(effort: str) -> str:
    """Map automation effort to CSS badge class."""
    from constants import AUTO_EFFORT_MAP
    return AUTO_EFFORT_MAP.get(effort.lower(), "badge-medium")


def get_root_cause_badge_class(category: str) -> str:
    """Map root cause category to CSS badge class."""
    from constants import ROOT_CAUSE_BADGE_MAP
    return ROOT_CAUSE_BADGE_MAP.get(category.lower(), "badge-medium")


def get_regression_risk_class(risk: str) -> str:
    """Map regression risk to CSS badge class."""
    from constants import REGRESSION_RISK_MAP
    return REGRESSION_RISK_MAP.get(risk.lower(), "badge-medium")


def get_trend_badge_class(trend: str) -> str:
    """Map trend direction to CSS badge class."""
    from constants import TREND_BADGE_MAP
    return TREND_BADGE_MAP.get(trend.lower(), "badge-medium")


def get_sprint_score_class(score: int) -> str:
    """Return CSS class for sprint health score colouring."""
    if score >= 80:
        return "score-good"
    elif score >= 60:
        return "score-warning"
    return "score-risk"


def get_compliance_score_class(score: int) -> str:
    """Return CSS class for compliance score colouring."""
    if score >= 80:
        return "score-good"
    elif score >= 50:
        return "score-warning"
    return "score-risk"


def get_issue_type_badge_class(issue_type: str) -> str:
    """Map issue type to CSS badge class."""
    from constants import ISSUE_TYPE_BADGE_MAP
    return ISSUE_TYPE_BADGE_MAP.get(issue_type.lower(), "badge-medium")


def get_ambiguity_score_class(score: int) -> str:
    """Return CSS class for ambiguity score colouring (lower is better)."""
    if score <= 20:
        return "score-good"
    elif score <= 60:
        return "score-warning"
    return "score-risk"


def get_invest_badge_class(status: str) -> str:
    """Map INVEST status to CSS badge class.

    Handles plain-text outputs like "PASS", "PARTIAL", "FAIL"
    and also legacy emoji-prefixed outputs for backward compatibility.
    """
    from constants import INVEST_BADGE_MAP
    # Extract the status word before any explanation/em-dash
    key = status.split("--")[0].strip().lower()
    # Strip any emoji prefixes for backward compatibility
    key = key.replace("\u2705", "").replace("\u26a0\ufe0f", "").replace("\u274c", "").strip()
    return INVEST_BADGE_MAP.get(key, "badge-medium")


def get_perf_framework_badge_class(framework: str) -> str:
    """Map performance framework to CSS badge class."""
    from constants import PERF_FRAMEWORK_BADGE_MAP
    return PERF_FRAMEWORK_BADGE_MAP.get(framework.lower(), "badge-medium")


def get_load_profile_badge_class(profile: str) -> str:
    """Map load profile type to CSS badge class."""
    from constants import LOAD_PROFILE_BADGE_MAP
    return LOAD_PROFILE_BADGE_MAP.get(profile.lower(), "badge-medium")


def get_criticality_badge_class(criticality: str) -> str:
    """Map criticality to CSS badge class."""
    from constants import CRITICALITY_BADGE_MAP
    return CRITICALITY_BADGE_MAP.get(criticality.lower(), "badge-medium")


def create_performance_zip(suite) -> bytes:
    """Package a PerformanceTestSuite into a downloadable ZIP archive (bytes)."""
    import io
    import zipfile

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Main script
        zf.writestr(suite.script_file_name, suite.script_code)

        # Execution plan
        plan_lines = [
            "# Performance Test Execution Plan",
            "",
            f"**Framework:** {suite.framework}",
            f"**Run Command:** `{suite.run_command}`",
            "",
            "## Setup",
            "",
        ]
        for cmd in suite.setup_instructions:
            plan_lines.append(f"```bash\n{cmd}\n```")
        plan_lines.extend([
            "",
            "## Execution Plan",
            "",
        ])
        for step in suite.execution_plan:
            plan_lines.append(f"- {step}")
        if suite.design_notes:
            plan_lines.extend(["", "## Design Notes", "", suite.design_notes])
        zf.writestr("execution-plan.md", "\n".join(plan_lines))

        # README
        readme_lines = [
            f"# {suite.framework} Performance Test Suite",
            "",
            "## Quick Start",
            "",
        ]
        for cmd in suite.setup_instructions:
            readme_lines.append(f"```bash\n{cmd}\n```")
        readme_lines.extend([
            "",
            "## Run",
            "",
            f"```bash\n{suite.run_command}\n```",
            "",
        ])
        if suite.design_notes:
            readme_lines.extend(["## Design Notes", "", suite.design_notes, ""])
        zf.writestr("README.md", "\n".join(readme_lines))

    buffer.seek(0)
    return buffer.read()


def get_test_type_badge_class(test_type: str) -> str:
    """Map security test type to CSS badge class."""
    from constants import TEST_TYPE_BADGE_MAP
    return TEST_TYPE_BADGE_MAP.get(test_type.lower(), "badge-medium")


def get_owasp_status_badge_class(status: str) -> str:
    """Map OWASP coverage status to CSS badge class."""
    from constants import OWASP_STATUS_BADGE_MAP
    return OWASP_STATUS_BADGE_MAP.get(status.lower(), "badge-medium")


def get_risk_level_badge_class(risk: str) -> str:
    """Map security risk level to CSS badge class."""
    from constants import RISK_LEVEL_BADGE_MAP
    return RISK_LEVEL_BADGE_MAP.get(risk.lower(), "badge-medium")


def get_security_score_class(score: int) -> str:
    """Return CSS class for security coverage score (higher is better)."""
    if score >= 80:
        return "score-good"
    elif score >= 50:
        return "score-warning"
    return "score-risk"


def export_security_to_excel(report) -> bytes:
    """Export a SecurityReport to a styled multi-sheet Excel workbook."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # ── Sheet 1: Security Test Cases ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Security Tests"

    headers = [
        "ID", "OWASP ID", "Category", "Title", "Description",
        "Steps", "Sample Payloads", "Severity", "Remediation", "Tool Hint", "Test Type"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for idx, tc in enumerate(report.test_cases, 2):
        row = [
            tc.id,
            tc.owasp_id,
            tc.owasp_category,
            tc.title,
            tc.description,
            "\n".join(f"{i+1}. {s}" for i, s in enumerate(tc.steps)),
            "\n".join(f"• {p}" for p in tc.sample_payloads),
            tc.severity,
            tc.remediation,
            tc.tool_hint,
            tc.test_type,
        ]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 8:  # Severity
                if val == "Critical":
                    cell.font = Font(color="DC2626", bold=True)
                elif val == "High":
                    cell.font = Font(color="D97706", bold=True)
                elif val == "Medium":
                    cell.font = Font(color="2563EB", bold=True)
                elif val == "Low":
                    cell.font = Font(color="059669", bold=True)

    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws1.column_dimensions[column].width = adjusted_width

    # ── Sheet 2: Payload Library ─────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Payload Library")
    pl_headers = ["Test ID", "OWASP ID", "Category", "Severity", "Payload"]
    for col, h in enumerate(pl_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for tc in report.test_cases:
        for payload in tc.sample_payloads:
            ws2.cell(row=row_idx, column=1, value=tc.id).border = thin_border
            ws2.cell(row=row_idx, column=2, value=tc.owasp_id).border = thin_border
            ws2.cell(row=row_idx, column=3, value=tc.owasp_category).border = thin_border
            ws2.cell(row=row_idx, column=4, value=tc.severity).border = thin_border
            ws2.cell(row=row_idx, column=5, value=payload).border = thin_border
            row_idx += 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 80

    # ── Sheet 3: OWASP Coverage ──────────────────────────────────────────────
    ws3 = wb.create_sheet(title="OWASP Coverage")
    cov_headers = ["OWASP ID", "Category", "Test Count", "Status"]
    for col, h in enumerate(cov_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, cov in enumerate(report.summary.owasp_coverage, 2):
        ws3.cell(row=idx, column=1, value=cov.owasp_id).border = thin_border
        ws3.cell(row=idx, column=2, value=cov.category).border = thin_border
        ws3.cell(row=idx, column=3, value=cov.test_count).border = thin_border
        ws3.cell(row=idx, column=4, value=cov.status).border = thin_border

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 15

    # ── Sheet 4: Summary ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet(title="Summary")
    summary = report.summary
    summary_rows = [
        ["Total Tests", summary.total_tests],
        ["Critical Count", summary.critical_count],
        ["High Count", summary.high_count],
        ["Medium Count", summary.medium_count],
        ["Low Count", summary.low_count],
        ["Coverage Score", summary.coverage_score],
        ["Overall Risk", summary.overall_risk],
    ]
    for idx, (k, v) in enumerate(summary_rows, 1):
        ws4.cell(row=idx, column=1, value=k).border = thin_border
        ws4.cell(row=idx, column=2, value=v).border = thin_border

    start_row = len(summary_rows) + 2
    ws4.cell(row=start_row, column=1, value="Key Findings").font = Font(bold=True, color="DC2626")
    for i, finding in enumerate(summary.key_findings, start_row + 1):
        ws4.cell(row=i, column=1, value=finding).border = thin_border
        ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    rec_start = start_row + len(summary.key_findings) + 2
    ws4.cell(row=rec_start, column=1, value="Recommendations").font = Font(bold=True, color="2563EB")
    for i, rec in enumerate(summary.recommendations, rec_start + 1):
        ws4.cell(row=i, column=1, value=rec).border = thin_border
        ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_test_cases_to_excel(result) -> bytes:
    """Export a TestCaseList to a styled Excel workbook (bytes)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd

    wb = Workbook()

    # ── Sheet 1: Test Cases ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Test Cases"

    headers = [
        "ID", "Title", "Category", "Priority", "Pre-conditions",
        "Steps", "Expected Result", "Test Data", "BDD Scenario",
        "Auto Feasibility", "Auto Effort", "Tags", "Traceability"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for idx, tc in enumerate(result.test_cases, 2):
        row = [
            tc.id,
            tc.title,
            tc.category,
            tc.priority,
            tc.pre_conditions,
            "\n".join(f"{i+1}. {s}" for i, s in enumerate(tc.steps)),
            tc.expected_result,
            tc.test_data or "",
            tc.bdd_scenario or "",
            tc.automation_feasibility,
            tc.automation_effort,
            ", ".join(tc.tags),
            tc.traceability,
        ]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Color-code priority
            if col == 4:  # Priority column
                if val == "High":
                    cell.font = Font(color="DC2626", bold=True)
                elif val == "Medium":
                    cell.font = Font(color="D97706", bold=True)
                elif val == "Low":
                    cell.font = Font(color="059669", bold=True)

    # Auto-width columns
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws1.column_dimensions[column].width = adjusted_width

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Summary")
    summary = result.summary

    summary_headers = ["Metric", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = [
        ["Total Generated", summary.total_generated],
        ["Automation Coverage Potential", summary.automation_coverage_potential],
    ]
    for cat in summary.category_breakdown:
        summary_rows.append([f"Category: {cat.category}", cat.count])
    for prio in summary.priority_breakdown:
        summary_rows.append([f"Priority: {prio.priority}", prio.count])

    for idx, (k, v) in enumerate(summary_rows, 2):
        ws2.cell(row=idx, column=1, value=k).border = thin_border
        ws2.cell(row=idx, column=2, value=v).border = thin_border

    # Coverage gaps
    start_row = len(summary_rows) + 3
    ws2.cell(row=start_row, column=1, value="Coverage Gaps").font = Font(bold=True, color="DC2626")
    for i, gap in enumerate(summary.coverage_gaps, start_row + 1):
        ws2.cell(row=i, column=1, value=gap).border = thin_border
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    # Recommendations
    rec_start = start_row + len(summary.coverage_gaps) + 2
    ws2.cell(row=rec_start, column=1, value="Recommendations").font = Font(bold=True, color="2563EB")
    for i, rec in enumerate(summary.recommendations, rec_start + 1):
        ws2.cell(row=i, column=1, value=rec).border = thin_border
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 60

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def render_page_header(title: str, desc: str, icon: str = "") -> None:
    """Render the breadcrumb + page title header used on every page."""
    import streamlit as st
    icon_html = f'<span class="page-icon">{icon}</span>' if icon else ""
    st.markdown(
        f'<div class="breadcrumb">QA-Genius &nbsp;&gt;&nbsp; {title}</div>\n'
        f'<div class="page-header">\n'
        f'    <h1>{icon_html} {title}</h1>\n'
        f'    <p>{desc}</p>\n'
        f'</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# ERROR CLASSES
# ══════════════════════════════════════════════════════════════════════════════

class TransientLLMError(Exception):
    """Errors that may succeed on retry (network, rate limit, timeout)."""
    pass


class PermanentLLMError(Exception):
    """Errors that won't succeed on retry (bad request, auth, content policy)."""
    pass


# ══════════════════════════════════════════════════════════════════════════════
# RESILIENT LLM CALLER
# ══════════════════════════════════════════════════════════════════════════════

def _classify_llm_error(e: Exception) -> Exception:
    """Classify an LLM exception as transient or permanent."""
    error_str = str(e).lower()

    if any(kw in error_str for kw in ["timeout", "timed out", "connection", "network"]):
        logger.warning(f"Transient network error: {type(e).__name__}")
        raise TransientLLMError(f"Network issue: {e}") from e

    if any(kw in error_str for kw in ["rate limit", "429", "too many requests", "quota"]):
        logger.warning(f"Rate limit hit: {type(e).__name__}")
        raise TransientLLMError(f"Rate limited: {e}") from e

    if any(kw in error_str for kw in ["unauthorized", "401", "403", "invalid api key", "authentication"]):
        logger.error(f"Auth error: {type(e).__name__}")
        raise PermanentLLMError(f"Authentication failed: {e}") from e

    if any(kw in error_str for kw in ["content_policy", "safety", "harmful", "blocked", "content filter"]):
        logger.warning(f"Content policy: {type(e).__name__}")
        raise PermanentLLMError(f"Content policy block: {e}") from e

    # Unknown — treat as transient and retry
    logger.warning(f"Unknown LLM error (will retry): {type(e).__name__}: {e}")
    raise TransientLLMError(f"Unexpected error: {e}") from e


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception(lambda e: isinstance(e, TransientLLMError)),
    reraise=True,
)
def _call_llm_with_retry(llm, prompt: str) -> str:
    """Internal: invokes LLM with automatic retry on transient errors."""
    try:
        response = llm.invoke(prompt)
        return response.content
    except Exception as e:
        raise _classify_llm_error(e)


def call_llm_safely(prompt: str, max_tokens: int = 8192) -> Optional[str]:
    """
    Public: call the LLM with full error handling.
    Returns the response string on success, None on failure (with st.error already shown to user).
    """
    try:
        llm = load_model()
        return _call_llm_with_retry(llm, prompt)

    except TransientLLMError as e:
        logger.error(f"All retries exhausted: {e}")
        st.error(
            "The AI service is temporarily unavailable. "
            "Please wait a moment and try again."
        )
        return None

    except PermanentLLMError as e:
        error_str = str(e).lower()
        if "authentication" in error_str:
            st.error(
                "Configuration error: the AI service is not accessible. "
                "If this persists, please report it on GitHub."
            )
        elif "content policy" in error_str:
            st.error(
                "Your input was flagged by the AI service's content policy. "
                "Please rephrase your request without sensitive or restricted topics."
            )
        else:
            st.error(
                "The AI service rejected this request. "
                "Try simplifying or rephrasing your input."
            )
        return None

    except Exception as e:
        logger.error(f"Unexpected error in call_llm_safely: {type(e).__name__}", exc_info=True)
        st.error(
            "An unexpected error occurred. Please try again with different input. "
            "If this persists, please report it on GitHub."
        )
        return None


# ══════════════════════════════════════════════════════════════════════════════
# PYDANTIC-AWARE PARSER WITH FALLBACK
# ══════════════════════════════════════════════════════════════════════════════

T = TypeVar("T", bound=BaseModel)


def parse_llm_response(
    raw_response: str,
    schema: Type[T],
    tab_name: str = "this tab",
) -> Optional[T]:
    """
    Parse LLM response into a Pydantic schema with resilient fallback.
    Shows user-friendly errors if parsing fails.
    """
    if not raw_response or not raw_response.strip():
        logger.error(f"Empty response from LLM for {tab_name}")
        st.error(
            f"The AI returned an empty response for {tab_name}. "
            f"Please try again with different input."
        )
        return None

    # Stage 1: resilient JSON parsing
    parsed_json = parse_json_resilient(raw_response)

    if parsed_json is None:
        logger.error(
            f"JSON parsing failed for {tab_name}. Response preview: "
            f"{raw_response[:300]}..."
        )
        st.error(
            f"The AI returned a response that couldn't be parsed for {tab_name}. "
            f"This usually means the input contained complex characters that confused the model. "
            f"Try simplifying the input or rephrasing it."
        )
        return None

    # Stage 2: Pydantic validation
    try:
        return schema.model_validate(parsed_json)
    except ValidationError as e:
        logger.error(
            f"Pydantic validation failed for {tab_name}: {e.error_count()} errors. "
            f"First error: {e.errors()[0] if e.errors() else 'none'}"
        )

        first_error = e.errors()[0] if e.errors() else {}
        field_name = ".".join(str(x) for x in first_error.get("loc", ["unknown"]))

        st.error(
            f"The AI's response was structured incorrectly for {tab_name} "
            f"(issue with field: `{field_name}`). "
            f"Please try again — this usually resolves on retry. "
            f"If it persists, simplify your input."
        )
        return None

    except Exception as e:
        logger.error(f"Unexpected parsing error for {tab_name}: {type(e).__name__}", exc_info=True)
        st.error(
            f"An unexpected error occurred while processing the AI response. "
            f"Please try again."
        )
        return None


# ══════════════════════════════════════════════════════════════════════════════
# ERROR ACTIONS & LOGGING
# ══════════════════════════════════════════════════════════════════════════════

def show_error_actions(tab_name: str, error_context: str = ""):
    """Show user-friendly recovery actions after an error."""
    col1, col2 = st.columns(2)

    with col1:
        if st.button("🔄 Try Again", key=f"retry_{tab_name}"):
            st.rerun()

    with col2:
        report_url = (
            f"https://github.com/Yusuf-Hridoy/QA-Genius/issues/new"
            f"?title=Error+in+{tab_name}"
            f"&body=Context:%20{error_context[:200]}"
        )
        st.markdown(f"[🐛 Report on GitHub]({report_url})")


def log_error_event(tab_name: str, error_type: str, **context):
    """Log a structured error event."""
    payload = {
        "event": "error",
        "tab": tab_name,
        "error_type": error_type,
        "timestamp": time.time(),
        **context,
    }
    logger.error(json.dumps(payload))


def log_success_event(tab_name: str, duration_ms: int, **context):
    """Log a successful generation."""
    payload = {
        "event": "success",
        "tab": tab_name,
        "duration_ms": duration_ms,
        "timestamp": time.time(),
        **context,
    }
    logger.info(json.dumps(payload))


def increment_error_count(tab_name: str, error_type: str):
    """Track error stats in session state for debugging."""
    if "error_log" not in st.session_state:
        st.session_state.error_log = []

    st.session_state.error_log.append({
        "tab": tab_name,
        "error_type": error_type,
        "timestamp": time.time(),
    })

    # Keep last 50 only
    st.session_state.error_log = st.session_state.error_log[-50:]


def create_automation_zip(script) -> bytes:
    """Package an AutomationScript into a downloadable ZIP archive (bytes)."""
    import io
    import zipfile

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        files_to_add = [
            (script.test_file_name, script.test_code),
        ]
        if script.page_object_file_name and script.page_object_code:
            files_to_add.append((script.page_object_file_name, script.page_object_code))
        if script.conftest_file_name and script.conftest_code:
            files_to_add.append((script.conftest_file_name, script.conftest_code))
        if script.config_file_name and script.config_code:
            files_to_add.append((script.config_file_name, script.config_code))
        if script.requirements_txt:
            files_to_add.append(("requirements.txt", script.requirements_txt))

        for filepath, content in files_to_add:
            zf.writestr(filepath, content)

        # Add a README
        readme_lines = [
            f"# {script.framework} Automation Project",
            "",
            "## Setup",
            "",
        ]
        for cmd in script.setup_instructions:
            readme_lines.append(f"```bash\n{cmd}\n```")
        readme_lines.extend([
            "",
            "## Run",
            "",
            f"```bash\n{script.execution_command}\n```",
            "",
        ])
        if script.design_notes:
            readme_lines.extend(["## Design Notes", "", script.design_notes, ""])
        zf.writestr("README.md", "\n".join(readme_lines))

    buffer.seek(0)
    return buffer.read()
